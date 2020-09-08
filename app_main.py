# -*- coding: utf-8 -*- 
# @Time : 2020/9/4 16:15 
# @Author : u4f55u6770 
# @contact: hejie@skyroam.com
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Series, Reference, LineChart, ScatterChart
from openpyxl.chart.axis import DateAxis

base_path = './source_files/'
# 文件名
file_name = 'source_file.xlsx'
# 表格名
sheet_name = 'Sheet1'


def histogram():
    # 查询参数
    bar_dict = {'第一轮': {'start_date': '2020-08-01 00:00:00', 'end_date': '2020-08-10 23:59:59'},
                '第二轮': {'start_date': '2020-08-11 00:00:00', 'end_date': '2020-08-20 23:59:59'},
                '第三轮': {'start_date': '2020-08-21 00:00:00', 'end_date': '2020-08-31 23:59:59'},
                '第四轮': {'start_date': '2020-09-01 00:00:00', 'end_date': '2020-09-04 23:59:59'}}
    # 过滤数据
    turn_filter(bar_dict)


def turn_filter(bar_dict):
    file_path = base_path + file_name
    if file_path:
        # 用pandas读取Excel,获取指定sheet的里的数据
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        # 创建时间筛选数据,根据数据生成多个条件,key为tag显示,len()为需要展示的数据
        rows = [('轮次', '缺陷数量')]
        char_index = 1
        for key in bar_dict.keys():
            start_date = bar_dict[key]['start_date']  # 开始时间
            end_date = bar_dict[key]['end_date']  # 结束时间
            # 筛选列表数据
            data = df[(df['创建日期'] >= start_date) & (df['创建日期'] <= end_date)]
            rows.append((key, len(data)))
            # print(key, start_date, end_date, len(data))
            char_index = char_index + 1

        # 打开文件，创建新的sheet
        wb = load_workbook(file_path)
        ws = wb.create_sheet('缺陷轮次统计')
        # 添加数据
        for row in rows:
            ws.append(row)
        histogram_chart = BarChart()
        histogram_chart.type = "col"
        histogram_chart.style = 10
        histogram_chart.title = "缺陷轮次统计"
        histogram_chart.y_axis.title = '数量'

        data = Reference(ws, min_col=2, min_row=1, max_row=7, max_col=2)
        cats = Reference(ws, min_col=1, min_row=2, max_row=7)
        histogram_chart.add_data(data, titles_from_data=True)
        histogram_chart.set_categories(cats)
        histogram_chart.shape = 4
        # 设置图表位置
        ws.add_chart(histogram_chart, "A{0}".format(char_index + 6))

        wb.save(file_path)


def line_graph():
    """
    生成折线图
    :return:
    """
    # 取某时间段内每天的数据做折线图
    start_date = '2020-08-10 00:00:00'
    end_date = '2020-08-30 23:59:59'

    # 读取数据
    file_path = base_path + file_name
    # 读取指定文件内的数据
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    # 获取时间段里的数据
    filter_data = df[(df['创建日期'] >= start_date) & (df['创建日期'] <= end_date)]
    # 准备格式化时间，做分组统计
    # filter_data['创建日期'] = [datetime.strftime(x, '%Y-%m-%d') for x in filter_data['创建日期']]
    # filter_data['创建日期'] = pd.to_datetime(filter_data['创建日期'])
    count_data = filter_data.groupby([filter_data['创建日期']], as_index=False)['ID'].count()
    # 转换时间格式,并重新组装数据
    rows = [('日期', '缺陷数量')]
    char_index = 1
    for row in count_data.itertuples():
        date_str = getattr(row, '创建日期')
        # print(date_str.strftime('%m月%d日'), getattr(row, 'ID'))
        rows.append((date_str.strftime('%m月%d日'), getattr(row, 'ID')))
        char_index = char_index + 1
    # 打开文件，创建新的sheet
    wb = load_workbook(file_path)
    ws = wb.create_sheet('每日缺陷曲线')
    # 添加数据
    for row in rows:
        ws.append(row)

    # 准备画折线图
    # Chart with date axis
    chart = LineChart()
    chart.title = "每日缺陷曲线"
    chart.style = 12
    # chart.y_axis.title = "Size"
    chart.y_axis.crossAx = 500
    chart.x_axis = DateAxis(crossAx=100)
    # chart.x_axis.number_format = 'd-mmm'
    # chart.x_axis.majorTimeUnit = "days"
    # chart.x_axis.title = "Date"
    # 图像的数据 起始行、起始列、终止行、终止列
    data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=char_index)
    chart.add_data(data, titles_from_data=True)
    dates = Reference(ws, min_col=1, min_row=2, max_row=char_index)
    chart.set_categories(dates)
    # 将图表添加到 sheet中
    ws.add_chart(chart, "A{0}".format(char_index + 6))

    wb.save(file_path)


def scatter():
    """
    根据工作项类型统计关闭数据
    :return:
    """
    # 工作项类型
    work_type = 'Bug'
    # 读取数据
    file_path = base_path + file_name
    # 读取指定文件内的数据
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    # 获取指定类型的数据
    filter_data = df[(df['工作项类型'] == work_type)]

    # 闭环周期数据
    closed_loop = {}
    for row in filter_data.itertuples():
        end_date_str = getattr(row, '关闭日期')
        if not pd.isna(end_date_str):
            start_date_str = getattr(row, '创建日期')
            # 闭环天数
            closed_day = (end_date_str - start_date_str).days
            if closed_day in closed_loop.keys():
                closed_num = closed_loop.get(closed_day)
                closed_loop[closed_day] = closed_num + 1
            else:
                closed_loop[closed_day] = 1
            # print(start_date_str, end_date_str, closed_day)
    # print(closed_loop)

    # 数据写入excel
    # 打开文件，创建新的sheet
    wb = load_workbook(file_path)
    ws = wb.create_sheet('{0}闭环统计'.format(work_type))
    # ws.append({'A': '闭环周期', 'B': '缺陷数'})
    # 排序
    closed_loop_rows = [['闭环周期', '缺陷数']]
    for key in sorted(closed_loop.keys(), reverse=False):
        # ws.append({'A': key, 'B': closed_loop[key]})
        closed_loop_rows.append([key, closed_loop[key]])
    # 写数据
    for row in closed_loop_rows:
        ws.append(row)
    # 设置图表
    row_len = len(closed_loop_rows)
    sc = ScatterChart()
    values = Reference(ws, min_col=1, min_row=1, max_row=row_len)
    x_values = Reference(ws, min_col=1, min_row=1, max_row=row_len)
    y_values = Reference(ws, min_col=2, min_row=1, max_row=row_len)
    series1 = Series(x_values, values, title_from_data=True)
    series2 = Series(y_values, values, title_from_data=True)
    sc.series.append(series1)
    sc.series.append(series2)
    # sc.title = "Default layout"
    sc.style = 12
    # sc.x_axis.title = 'Size'
    # sc.y_axis.title = 'Percentage'
    ws.add_chart(sc, "A{0}".format(row_len + 6))

    # 指定类型每日关闭数据
    closed_count_data = filter_data.groupby([filter_data['关闭日期']], as_index=False)['ID'].count()
    char_index = 1
    total = 0
    # 装换时间
    rows = [('日期', '缺陷数量')]
    for row in closed_count_data.itertuples():
        date_str = getattr(row, '关闭日期')
        num = getattr(row, 'ID')
        # print(date_str.strftime('%m月%d日'), num)
        total = total + num
        rows.append((date_str.strftime('%m月%d日'), num))
        char_index = char_index + 1
    rows.append(('总计', total))
    # print(rows)

    # 要标记位置 wk_sheet.cell(row=2,column=4,value='店铺编码')
    row_index = 1
    for row in rows:
        ws.cell(row=row_index, column=4, value=row[0])
        ws.cell(row=row_index, column=5, value=row[1])
        row_index = row_index + 1

    wb.save(file_path)


if __name__ == '__main__':
    # 生成直方图
    histogram()

    # 生成折线图
    line_graph()

    # 生成透析表
    scatter()
